import os
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import io
import hashlib
from tqdm import tqdm

def setup_driver(width=1280, height=720, headless=True):
    """设置Chrome浏览器驱动"""
    chrome_options = Options()
    if headless:
        chrome_options.add_argument("--headless")
    chrome_options.add_argument(f"--window-size={width},{height}")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    return driver

def get_image_hash(image_data):
    """计算图像数据的哈希值，用于去重"""
    return hashlib.md5(image_data).hexdigest()

def read_tiktok_links(file_path):
    """读取包含TikTok链接的文本文件"""
    with open(file_path, 'r') as file:
        links = [line.strip() for line in file if line.strip()]
    return links

def extract_username(url):
    """从TikTok URL中提取用户名"""
    return url.split('@')[1] if '@' in url else "未知用户"

def create_excel_with_screenshots(links_file, output_excel, width=1280, height=720):
    """创建包含截图和标题的Excel文件"""
    links = read_tiktok_links(links_file)
    
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "TikTok用户"
    
    # 设置表头
    ws['A1'] = "用户名"
    ws['B1'] = "截图"
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    
    # 行高设置
    ws.row_dimensions[1].height = 20
    
    # 设置驱动
    driver = setup_driver(width, height)
    
    # 创建图片保存目录
    temp_dir = "temp_screenshots"
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)
    
    # 用于去重的哈希集合
    image_hashes = set()
    
    # 使用tqdm显示进度
    print("开始处理TikTok链接...")
    row = 2
    
    for i, url in enumerate(tqdm(links, desc="正在处理TikTok链接")):
        try:
            # 获取用户名
            username = extract_username(url)
            
            # 访问链接
            driver.get(url)
            time.sleep(3)  # 等待页面加载
            
            # 截图
            screenshot = driver.get_screenshot_as_png()
            image_hash = get_image_hash(screenshot)
            
            # 检查图片是否重复
            if image_hash not in image_hashes:
                image_hashes.add(image_hash)
                
                # 保存截图到临时文件
                temp_img_path = os.path.join(temp_dir, f"{username}.png")
                with open(temp_img_path, 'wb') as f:
                    f.write(screenshot)
                
                # 调整图片大小以适应Excel单元格
                pil_img = PILImage.open(io.BytesIO(screenshot))
                new_width = 400
                ratio = new_width / pil_img.width
                new_height = int(pil_img.height * ratio)
                pil_img = pil_img.resize((new_width, new_height))
                pil_img.save(temp_img_path)
                
                # 将数据写入Excel
                ws.cell(row=row, column=1).value = username
                img = Image(temp_img_path)
                
                # 调整行高以适应图片
                ws.row_dimensions[row].height = new_height * 0.75
                
                # 插入图片
                ws.add_image(img, f'B{row}')
                
                row += 1
            else:
                print(f"跳过重复图片: {username}")
                
        except Exception as e:
            print(f"处理链接 {url} 时出错: {str(e)}")
    
    # 保存Excel文件
    wb.save(output_excel)
    
    # 关闭浏览器
    driver.quit()
    
    # 清理临时文件
    for file in os.listdir(temp_dir):
        os.remove(os.path.join(temp_dir, file))
    os.rmdir(temp_dir)
    
    print(f"完成! Excel文件已保存为 {output_excel}")
    print(f"共处理 {len(links)} 个链接，成功添加 {row-2} 个条目")

if __name__ == "__main__":
    # 用户输入
    links_file = input("请输入包含TikTok链接的文本文件路径: ")
    output_excel = input("请输入输出Excel文件名 (默认: tiktok_screenshots.xlsx): ") or "tiktok_screenshots.xlsx"
    width = int(input("请输入截图宽度 (默认: 1280): ") or "1280")
    height = int(input("请输入截图高度 (默认: 720): ") or "720")
    
    create_excel_with_screenshots(links_file, output_excel, width, height)
